import requests
from bs4 import BeautifulSoup
import openpyxl
import time
from openpyxl.styles import NamedStyle, Font

def scrape_hotels():
    try:
        start_time = time.time()

        # Read the URL list from an Excel file
        excel_file_path = "hotel4.xlsx"
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        column_a = [cell.value for cell in worksheet["A"][1:]]

        data = []

        for url in column_a:
            response = requests.get(url)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')

                # Extract the data from the HTML source
                title = soup.title.text.strip()
                address_element = soup.select_one("span.hp_address_subtitle.js-hp_address_subtitle.jq_tooltip")
                address = address_element.text.strip() if address_element else ""

                # Separate address into components
                address_components = address.split(",")

                # Further split the postal code and city
                postal_code_and_city = address_components[1].strip().split()

                # Extract postal code and city
                postal_code = postal_code_and_city[0] if postal_code_and_city else ""
                city = ' '.join(postal_code_and_city[1:]) if len(postal_code_and_city) > 1 else ""
                state = address_components[2].strip() if len(address_components) > 2 else ""

                entry = {
                    "link": url,
                    "title": title,
                    "address": address,
                    "postalCode": postal_code,
                    "city": city,
                    "state": state,
                }

                data.append(entry)

        # Save the data to an Excel file
        output_workbook = openpyxl.Workbook()
        output_worksheet = output_workbook.active

        # Add headers
        output_worksheet.append(["Link", "Title", "Address", "Postal Code", "City", "State"])

        # Create a named style for the hyperlink
        hyperlink_style = NamedStyle(name='hyperlink', font=Font(color='0000FF', underline='single'))

        # Add data with hyperlinks
        for entry in data:
            row_data = list(entry.values())
            link_cell = row_data[0]

            # Add hyperlink to the link cell
            output_worksheet.cell(row=output_worksheet.max_row + 1, column=1).hyperlink = link_cell
            output_worksheet.cell(row=output_worksheet.max_row, column=1).value = link_cell
            output_worksheet.cell(row=output_worksheet.max_row, column=1).style = hyperlink_style

            # Add other data to the rest of the columns
            for col_num, data_value in enumerate(row_data[1:], start=2):
                output_worksheet.cell(row=output_worksheet.max_row, column=col_num, value=data_value)

        # Save the output workbook
        output_workbook.save("hotel4-address.xlsx")

        end_time = time.time()
        execution_time = end_time - start_time
        print(f"Script Execution Time: {execution_time} seconds")

    except Exception as e:
        print(e)

if __name__ == "__main__":
    scrape_hotels()
